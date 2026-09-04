import streamlit as st
from openpyxl import load_workbook
import pandas as pd
import os
import re
import tempfile
import io
import zipfile


# ============================================================
# STREAMLIT PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="MRF Automation",
    page_icon="📊",
    layout="wide"
)


# ============================================================
# PAGE TITLE
# ============================================================

st.title("📊 MRF Automation")
st.write("Upload the MRF template and From Solution Excel file to generate MRF files.")


# ============================================================
# FILE UPLOAD
# ============================================================

col1, col2 = st.columns(2)

with col1:

    template_uploaded = st.file_uploader(
        "Upload MRF Template",
        type=["xlsx"],
        key="template"
    )


with col2:

    from_uploaded = st.file_uploader(
        "Upload From Solution Data",
        type=["xlsx"],
        key="from_solution"
    )


st.divider()


# ============================================================
# MRF PROCESSING FUNCTION
# ============================================================

def create_mrf_files(template_file, from_file):

    created_files = []

    # --------------------------------------------------------
    # Create temporary folder
    # --------------------------------------------------------

    temp_dir = tempfile.mkdtemp()

    template_path = os.path.join(
        temp_dir,
        "MRF_Sample.xlsx"
    )

    from_path = os.path.join(
        temp_dir,
        "from_solution_file.xlsx"
    )

    output_folder = os.path.join(
        temp_dir,
        "MRF_Make_output_file"
    )

    os.makedirs(
        output_folder,
        exist_ok=True
    )

    # --------------------------------------------------------
    # Save uploaded files temporarily
    # --------------------------------------------------------

    with open(template_path, "wb") as f:

        f.write(
            template_file.getbuffer()
        )


    with open(from_path, "wb") as f:

        f.write(
            from_file.getbuffer()
        )


    # --------------------------------------------------------
    # Read Excel
    # --------------------------------------------------------

    df = pd.read_excel(
        from_path
    )


    # --------------------------------------------------------
    # Clean Site ID column
    # --------------------------------------------------------

    df.iloc[:, 0] = (
        df.iloc[:, 0]
        .fillna("")
        .astype(str)
        .str.strip()
    )


    site_ids = df.iloc[:, 0].unique()


    # ========================================================
    # FUNCTIONS
    # ========================================================

    def clean_list(data):

        return [

            str(v).strip()

            for v in data

            if str(v).strip() != ""

            and str(v).lower() != "nan"

        ]


    def safe_num(val):

        try:

            return int(float(val))

        except:

            return 0


    def safe_col(data, index):

        if len(data.columns) > index:

            return data.iloc[:, index].tolist()

        return []


    def check_model(value, models):

        text = str(value).upper()

        text = text.replace(" ", "")


        for model in models:

            if model in text:

                return True


        return False


    def check_band(value, band):

        text = str(value).upper()

        return band in text


    # ========================================================
    # PROCESS EACH SITE
    # ========================================================

    created_file_count = 0


    for site in site_ids:

        if not site:

            continue


        matched_rows = df[
            df.iloc[:, 0]
            .astype(str)
            .str.strip()
            ==
            str(site).strip()
        ]


        if matched_rows.empty:

            continue


        # ----------------------------------------------------
        # Load template
        # ----------------------------------------------------

        wb = load_workbook(
            template_path
        )

        ws = wb.active


        # ----------------------------------------------------
        # Basic values
        # ----------------------------------------------------

        val_P = (
            matched_rows.iloc[0, 15]
            if len(matched_rows.columns) > 15
            else ""
        )

        val_Q = (
            matched_rows.iloc[0, 16]
            if len(matched_rows.columns) > 16
            else ""
        )

        val_R = (
            matched_rows.iloc[0, 17]
            if len(matched_rows.columns) > 17
            else ""
        )

        val_S = (
            matched_rows.iloc[0, 18]
            if len(matched_rows.columns) > 18
            else ""
        )

        val_T = (
            matched_rows.iloc[0, 19]
            if len(matched_rows.columns) > 19
            else ""
        )


        # ----------------------------------------------------
        # Fill template
        # ----------------------------------------------------

        ws["G8"] = val_P
        ws["G9"] = val_Q
        ws["G10"] = site
        ws["G11"] = val_R

        ws["J8"] = val_T
        ws["J9"] = val_S
        ws["J10"] = val_T
        ws["J11"] = val_S


        # ----------------------------------------------------
        # Counters
        # ----------------------------------------------------

        count_2219_b8 = 0

        count_4415_4428_b3 = 0

        count_2219_2217_b1 = 0

        count_4480_4499_b1_b3 = 0


        # ====================================================
        # PROCESS MATCHED ROWS
        # ====================================================

        for _, row in matched_rows.iterrows():

            b8 = str(row.iloc[1]).upper()

            b1 = str(row.iloc[2]).upper()

            b3 = str(row.iloc[3]).upper()


            full_row = (

                b8
                +
                " "
                +
                b1
                +
                " "
                +
                b3

            )


            # ------------------------------------------------
            # 2219 B8
            # ------------------------------------------------

            if (

                check_model(
                    full_row,
                    ["2219"]
                )

                and

                check_band(
                    b8,
                    "B8"
                )

            ):

                count_2219_b8 += 1


            # ------------------------------------------------
            # 4415 / 4428 B3
            # ------------------------------------------------

            if (

                check_model(
                    full_row,
                    ["4415", "4428"]
                )

                and

                check_band(
                    b3,
                    "B3"
                )

            ):

                count_4415_4428_b3 += 1


            # ------------------------------------------------
            # 2219 / 2217 B1
            # ------------------------------------------------

            model_text = full_row.replace(
                " ",
                ""
            )


            if (

                (

                    "2219/2217" in model_text

                    or

                    "2217/2219" in model_text

                )

                and

                "B1" in b1

            ):

                count_2219_2217_b1 += 1


            # ------------------------------------------------
            # 4480 / 4499 B1 B3
            # ------------------------------------------------

            if (

                check_model(
                    full_row,
                    ["4480", "4499"]
                )

                and

                "B1" in full_row

                and

                "B3" in full_row

            ):

                count_4480_4499_b1_b3 += 1


        # ====================================================
        # WRITE COUNTS
        # ====================================================

        ws["G15"] = (

            "2219 B8"

            if count_2219_b8

            else ""

        )

        ws["I15"] = count_2219_b8


        ws["G16"] = (

            "4415/4428 B3"

            if count_4415_4428_b3

            else ""

        )

        ws["I16"] = count_4415_4428_b3


        ws["G17"] = (

            "2219/2217 B1"

            if count_2219_2217_b1

            else ""

        )

        ws["I17"] = count_2219_2217_b1


        ws["G18"] = (

            "4480/4499 B1 B3"

            if count_4480_4499_b1_b3

            else ""

        )

        ws["I18"] = count_4480_4499_b1_b3


        # ====================================================
        # 4480 / 4499
        # ====================================================

        if count_4480_4499_b1_b3:

            ws["G31"] = "Circular Power Connector"

            ws["I31"] = 1


            ws["G32"] = "Dual ERS heavy bracket"

            ws["I32"] = 1


        else:

            ws["G31"] = ""

            ws["I31"] = ""


            ws["G32"] = ""

            ws["I32"] = ""


        # ====================================================
        # GSM
        # ====================================================

        gsm_values = clean_list(
            safe_col(
                matched_rows,
                4
            )
        )


        gsm_total = 0


        for v in gsm_values:

            try:

                gsm_total += int(float(v))

            except:

                pass


        if gsm_total != 0:

            ws["G24"] = "GSM"

            ws["I24"] = gsm_total


            ws["G25"] = "Down Tilt H"

            ws["I25"] = gsm_total


            ws["G26"] = (
                "SR:RET Control Cable "
                "(3GPP / AISG) 5 m"
            )

            ws["I26"] = gsm_total


        else:

            ws["G24"] = ""

            ws["I24"] = ""

            ws["G25"] = ""

            ws["I25"] = ""

            ws["G26"] = ""

            ws["I26"] = ""


        # ====================================================
        # SFP
        # ====================================================

        i_values = clean_list(
            safe_col(
                matched_rows,
                8
            )
        )


        i_numeric = []


        for v in i_values:

            try:

                i_numeric.append(
                    int(float(v))
                )

            except:

                pass


        ws["G19"] = (

            "SFP"

            if i_numeric

            else ""

        )


        ws["I19"] = (

            sum(i_numeric)

            if i_numeric

            else 0

        )


        # ====================================================
        # RRU CONNECTOR
        # ====================================================

        j_values = clean_list(
            safe_col(
                matched_rows,
                9
            )
        )


        j_numeric = []


        for v in j_values:

            try:

                j_numeric.append(
                    int(float(v))
                )

            except:

                pass


        ws["G20"] = "RRU Connector"


        ws["I20"] = (

            sum(j_numeric)

            if j_numeric

            else 0

        )


        # ====================================================
        # 4.3 TO 4.3 JUMPER
        # ====================================================

        l_values = clean_list(
            safe_col(
                matched_rows,
                10
            )
        )


        l_numeric = []


        for v in l_values:

            try:

                l_numeric.append(
                    int(float(v))
                )

            except:

                pass


        ws["G21"] = "4.3 to 4.3 jumper"


        ws["I21"] = (

            sum(l_numeric)

            if l_numeric

            else 0

        )


        # ====================================================
        # 4.3 TO 7/16 JUMPER
        # ====================================================

        k_values = clean_list(
            safe_col(
                matched_rows,
                11
            )
        )


        k_numeric = []


        for v in k_values:

            try:

                k_numeric.append(
                    int(float(v))
                )

            except:

                pass


        ws["G22"] = "4.3 to 7/16 jumper"


        ws["I22"] = (

            sum(k_numeric)

            if k_numeric

            else 0

        )


        # ====================================================
        # RRU POWER CABLE
        # ====================================================

        h_values = clean_list(
            safe_col(
                matched_rows,
                12
            )
        )


        h_numeric = []


        for v in h_values:

            try:

                h_numeric.append(
                    int(float(v))
                )

            except:

                pass


        ws["G23"] = "RRU pw cable"


        ws["I23"] = (

            sum(h_numeric)

            if h_numeric

            else 0

        )


        # ====================================================
        # TOTAL QUANTITY
        # ====================================================

        extra_minus = (

            1

            if count_4480_4499_b1_b3

            else 0

        )


        total_quantity = (

            safe_num(ws["I15"].value)

            +

            safe_num(ws["I16"].value)

            +

            safe_num(ws["I17"].value)

            +

            safe_num(ws["I18"].value)

            -

            extra_minus

        )


        ws["I27"] = (

            total_quantity

            if total_quantity != 0

            else ""

        )


        ws["I28"] = (

            total_quantity

            if total_quantity != 0

            else ""

        )


        # ====================================================
        # HIDE / SHOW ROWS
        # ====================================================

        for r in range(15, 33):

            value = ws[f"I{r}"].value


            if value in [0, "0", None, ""]:

                ws.row_dimensions[r].hidden = True

            else:

                ws.row_dimensions[r].hidden = False


        # ====================================================
        # SUFFIX
        # ====================================================

        suffix = "4K"


        if val_R:

            r_text = str(
                val_R
            ).strip().upper()


            if "GPI" in r_text:

                suffix = "GPI"


            elif "CFVL" in r_text:

                suffix = "CFVL"


            else:

                suffix = r_text


        # ====================================================
        # SAFE SITE NAME
        # ====================================================

        safe_site = re.sub(

            r'[\\/*?:"<>|]',

            "_",

            str(site)

        )


        # ====================================================
        # OUTPUT FILE
        # ====================================================

        output_path = os.path.join(

            output_folder,

            f"MRF_{safe_site}_{suffix}.xlsx"

        )


        # ----------------------------------------------------
        # Save
        # ----------------------------------------------------

        wb.save(
            output_path
        )


        created_file_count += 1


        created_files.append(
            output_path
        )


    return created_files


# ============================================================
# GENERATE BUTTON
# ============================================================

if st.button(
    "🚀 Generate MRF Files",
    type="primary",
    use_container_width=True
):

    if template_uploaded is None:

        st.error(
            "❌ Please upload the MRF Template Excel file."
        )


    elif from_uploaded is None:

        st.error(
            "❌ Please upload the From Solution Data Excel file."
        )


    else:

        try:

            with st.spinner(
                "Generating MRF files... Please wait."
            ):

                created_files = create_mrf_files(
                    template_uploaded,
                    from_uploaded
                )


            # =================================================
            # RESULT
            # =================================================

            if not created_files:

                st.warning(
                    "⚠️ No MRF files were created."
                )


            else:

                st.success(
                    f"✅ MRF generation completed successfully! "
                    f"Total files created: {len(created_files)}"
                )


                st.divider()

                st.subheader(
                    "📁 Generated MRF Files"
                )


                # =================================================
                # INDIVIDUAL DOWNLOAD BUTTONS
                # =================================================

                for file_path in created_files:

                    file_name = os.path.basename(
                        file_path
                    )


                    with open(
                        file_path,
                        "rb"
                    ) as f:

                        file_data = f.read()


                    st.download_button(

                        label=f"⬇️ Download {file_name}",

                        data=file_data,

                        file_name=file_name,

                        mime=(
                            "application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet"
                        ),

                        use_container_width=True

                    )


                # =================================================
                # ZIP FILE
                # =================================================

                zip_buffer = io.BytesIO()


                with zipfile.ZipFile(
                    zip_buffer,
                    "w",
                    zipfile.ZIP_DEFLATED
                ) as zip_file:

                    for file_path in created_files:

                        zip_file.write(

                            file_path,

                            arcname=os.path.basename(
                                file_path
                            )

                        )


                zip_buffer.seek(0)


                st.divider()

                st.download_button(

                    label="📦 Download All MRF Files (ZIP)",

                    data=zip_buffer,

                    file_name="MRF_All_Files.zip",

                    mime="application/zip",

                    use_container_width=True

                )


        except Exception as e:

            st.error(
                "❌ An error occurred while generating the MRF files."
            )

            st.exception(e)