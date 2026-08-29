import streamlit as st
from openpyxl import load_workbook
import pandas as pd
import os
import re
import tempfile
import io
import zipfile


# ============================================================
# STREAMLIT PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="MRF Automation",
    page_icon="📊",
    layout="wide"
)

st.title("📊 MRF Automation")
st.write(
    "Upload the MRF Template, From Solution Data, and Item Code Rules file. "
    "The app will generate MRF files and automatically add item codes."
)


# ============================================================
# FILE UPLOAD
# ============================================================

col1, col2, col3 = st.columns(3)

with col1:
    template_uploaded = st.file_uploader(
        "Upload MRF Template",
        type=["xlsx"],
        key="template"
    )

with col2:
    from_uploaded = st.file_uploader(
        "Upload From Solution Data",
        type=["xlsx"],
        key="from_solution"
    )

with col3:
    rules_uploaded = st.file_uploader(
        "Upload Item Code Rules File",
        type=["xlsx"],
        key="rules"
    )

st.divider()


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_list(data):
    return [
        str(v).strip()
        for v in data
        if str(v).strip() != ""
        and str(v).lower() != "nan"
    ]


def safe_num(val):
    try:
        return int(float(val))
    except Exception:
        return 0


def safe_col(data, index):
    if len(data.columns) > index:
        return data.iloc[:, index].tolist()
    return []


def check_model(value, models):
    text = str(value).upper().replace(" ", "")
    return any(model in text for model in models)


def check_band(value, band):
    return band in str(value).upper()


def norm(v):
    if pd.isna(v):
        return ""

    return str(v).strip().upper().replace(" ", "")


# ============================================================
# LOAD ITEM CODE RULES
# ============================================================

def load_item_code_rules(rules_path):
    df_rules = pd.read_excel(
        rules_path,
        sheet_name="RULES",
        dtype=object
    ).fillna("")

    df_rules = df_rules.infer_objects(copy=False)

    exact_map = {}
    smart_map = {}

    for i in range(len(df_rules)):
        key = norm(df_rules.iloc[i, 0])

        if not key:
            continue

        values = df_rules.iloc[i, 1:6].tolist()
        exact_map[key] = values

        if "B1" in key or "B3" in key or "B8" in key:
            smart_map[key] = values

    return exact_map, smart_map


def smart_match(g, smart_map):
    for key, values in smart_map.items():
        rule = key.replace(" ", "")

        if "B1" in rule:
            b_part = "B1"
            item_part = rule.replace("B1", "")
        elif "B3" in rule:
            b_part = "B3"
            item_part = rule.replace("B3", "")
        elif "B8" in rule:
            b_part = "B8"
            item_part = rule.replace("B8", "")
        else:
            continue

        if b_part not in g:
            continue

        codes = item_part.split("/")

        for code in codes:
            if code and code in g:
                return values

    return None


def write_rule_values(ws, row, values, fallback_item=None):
    ws[f"C{row}"] = values[0] if len(values) > 0 else ""
    ws[f"D{row}"] = values[1] if len(values) > 1 else ""
    ws[f"E{row}"] = values[2] if len(values) > 2 else ""
    ws[f"F{row}"] = values[3] if len(values) > 3 else ""

    if len(values) > 4:
        ws[f"G{row}"] = values[4]
    elif fallback_item is not None:
        ws[f"G{row}"] = fallback_item


def apply_item_codes(wb, exact_map, smart_map):
    updated = 0

    for ws in wb.worksheets:

        for r in range(1, ws.max_row + 1):
            g = norm(ws[f"G{r}"].value)

            if not g:
                continue

            values = exact_map.get(g)

            if values is None:
                values = smart_match(g, smart_map)

            if values is not None:
                write_rule_values(
                    ws,
                    r,
                    values,
                    ws[f"G{r}"].value
                )
                updated += 1

        # Extra rules created by 4480 / 4499 logic
        for row, item_name in [
            (31, "Circular Power Connector"),
            (32, "Dual ERS heavy bracket")
        ]:
            if ws[f"I{row}"].value == 1:
                key = norm(item_name)

                if key in exact_map:
                    write_rule_values(
                        ws,
                        row,
                        exact_map[key],
                        item_name
                    )
                    updated += 1
                else:
                    ws[f"G{row}"] = item_name

    return updated


# ============================================================
# MAIN MRF PROCESSING FUNCTION
# ============================================================

def create_mrf_files(template_file, from_file, rules_file):

    created_files = []
    temp_dir = tempfile.mkdtemp()

    template_path = os.path.join(temp_dir, "MRF_Sample.xlsx")
    from_path = os.path.join(temp_dir, "from_solution_file.xlsx")
    rules_path = os.path.join(temp_dir, "item_code_rules.xlsx")

    output_folder = os.path.join(temp_dir, "MRF_Make_output_file")
    os.makedirs(output_folder, exist_ok=True)

    # Save uploaded files
    with open(template_path, "wb") as f:
        f.write(template_file.getbuffer())

    with open(from_path, "wb") as f:
        f.write(from_file.getbuffer())

    with open(rules_path, "wb") as f:
        f.write(rules_file.getbuffer())

    # Read source and rules
    df = pd.read_excel(from_path)
    df.iloc[:, 0] = (
        df.iloc[:, 0]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    site_ids = df.iloc[:, 0].unique()
    exact_map, smart_map = load_item_code_rules(rules_path)

    # ========================================================
    # PROCESS EACH SITE
    # ========================================================

    for site in site_ids:

        if not site:
            continue

        matched_rows = df[
            df.iloc[:, 0]
            .astype(str)
            .str.strip()
            == str(site).strip()
        ]

        if matched_rows.empty:
            continue

        wb = load_workbook(template_path)
        ws = wb.active

        # ----------------------------------------------------
        # BASIC VALUES
        # ----------------------------------------------------

        val_P = matched_rows.iloc[0, 15] if len(matched_rows.columns) > 15 else ""
        val_Q = matched_rows.iloc[0, 16] if len(matched_rows.columns) > 16 else ""
        val_R = matched_rows.iloc[0, 17] if len(matched_rows.columns) > 17 else ""
        val_S = matched_rows.iloc[0, 18] if len(matched_rows.columns) > 18 else ""
        val_T = matched_rows.iloc[0, 19] if len(matched_rows.columns) > 19 else ""

        ws["G8"] = val_P
        ws["G9"] = val_Q
        ws["G10"] = site
        ws["G11"] = val_R

        ws["J8"] = val_T
        ws["J9"] = val_S
        ws["J10"] = val_T
        ws["J11"] = val_S

        # ----------------------------------------------------
        # COUNTERS
        # ----------------------------------------------------

        count_2219_b8 = 0
        count_4415_4428_b3 = 0
        count_2219_2217_b1 = 0
        count_4480_4499_b1_b3 = 0

        for _, row in matched_rows.iterrows():

            b8 = str(row.iloc[1]).upper()
            b1 = str(row.iloc[2]).upper()
            b3 = str(row.iloc[3]).upper()

            full_row = f"{b8} {b1} {b3}"

            # 2219 B8
            if check_model(full_row, ["2219"]) and check_band(b8, "B8"):
                count_2219_b8 += 1

            # 4415 / 4428 B3
            if check_model(full_row, ["4415", "4428"]) and check_band(b3, "B3"):
                count_4415_4428_b3 += 1

            # 2219 / 2217 B1
            model_text = full_row.replace(" ", "")

            if (
                ("2219/2217" in model_text or "2217/2219" in model_text)
                and "B1" in b1
            ):
                count_2219_2217_b1 += 1

            # 4480 / 4499 B1 B3
            if (
                check_model(full_row, ["4480", "4499"])
                and "B1" in full_row
                and "B3" in full_row
            ):
                count_4480_4499_b1_b3 += 1

        # ----------------------------------------------------
        # WRITE COUNTS
        # ----------------------------------------------------

        ws["G15"] = "2219 B8" if count_2219_b8 else ""
        ws["I15"] = count_2219_b8

        ws["G16"] = "4415/4428 B3" if count_4415_4428_b3 else ""
        ws["I16"] = count_4415_4428_b3

        ws["G17"] = "2219/2217 B1" if count_2219_2217_b1 else ""
        ws["I17"] = count_2219_2217_b1

        ws["G18"] = "4480/4499 B1 B3" if count_4480_4499_b1_b3 else ""
        ws["I18"] = count_4480_4499_b1_b3

        # ----------------------------------------------------
        # 4480 / 4499 EXTRA ITEMS
        # ----------------------------------------------------

        if count_4480_4499_b1_b3:
            ws["G31"] = "Circular Power Connector"
            ws["I31"] = 1

            ws["G32"] = "Dual ERS heavy bracket"
            ws["I32"] = 1
        else:
            ws["G31"] = ""
            ws["I31"] = ""
            ws["G32"] = ""
            ws["I32"] = ""

        # ----------------------------------------------------
        # GSM
        # ----------------------------------------------------

        gsm_total = sum(
            safe_num(v)
            for v in clean_list(safe_col(matched_rows, 4))
        )

        if gsm_total != 0:
            ws["G24"] = "GSM"
            ws["I24"] = gsm_total

            ws["G25"] = "Down Tilt H"
            ws["I25"] = gsm_total

            ws["G26"] = "SR:RET Control Cable (3GPP / AISG) 5 m"
            ws["I26"] = gsm_total
        else:
            for cell in ["G24", "I24", "G25", "I25", "G26", "I26"]:
                ws[cell] = ""

        # ----------------------------------------------------
        # SUM COLUMNS
        # ----------------------------------------------------

        item_rules = [
            (8,  "G19", "I19", "SFP"),
            (9,  "G20", "I20", "RRU Connector"),
            (10, "G21", "I21", "4.3 to 4.3 jumper"),
            (11, "G22", "I22", "4.3 to 7/16 jumper"),
            (12, "G23", "I23", "RRU pw cable"),
        ]

        for col_index, g_cell, i_cell, item_name in item_rules:
            total = sum(
                safe_num(v)
                for v in clean_list(safe_col(matched_rows, col_index))
            )
            ws[g_cell] = item_name
            ws[i_cell] = total if total else 0

        # ----------------------------------------------------
        # TOTAL QUANTITY
        # ----------------------------------------------------

        extra_minus = 1 if count_4480_4499_b1_b3 else 0

        total_quantity = (
            safe_num(ws["I15"].value)
            + safe_num(ws["I16"].value)
            + safe_num(ws["I17"].value)
            + safe_num(ws["I18"].value)
            - extra_minus
        )

        ws["I27"] = total_quantity if total_quantity != 0 else ""
        ws["I28"] = total_quantity if total_quantity != 0 else ""

        # ----------------------------------------------------
        # HIDE / SHOW ROWS
        # ----------------------------------------------------

        for r in range(15, 33):
            value = ws[f"I{r}"].value
            ws.row_dimensions[r].hidden = value in [0, "0", None, ""]

        # ----------------------------------------------------
        # APPLY ITEM CODES (SECOND SCRIPT)
        # ----------------------------------------------------

        apply_item_codes(wb, exact_map, smart_map)

        # ----------------------------------------------------
        # SUFFIX
        # ----------------------------------------------------

        suffix = "4K"

        if val_R:
            r_text = str(val_R).strip().upper()

            if "GPI" in r_text:
                suffix = "GPI"
            elif "CFVL" in r_text:
                suffix = "CFVL"
            else:
                suffix = r_text

        safe_site = re.sub(
            r'[\\/*?:"<>|]',
            "_",
            str(site)
        )

        output_path = os.path.join(
            output_folder,
            f"MRF_{safe_site}_{suffix}.xlsx"
        )

        wb.save(output_path)
        wb.close()

        created_files.append(output_path)

    return created_files


# ============================================================
# GENERATE BUTTON
# ============================================================

if st.button(
    "🚀 Generate MRF Files + Add Item Codes",
    type="primary",
    use_container_width=True
):

    if template_uploaded is None:
        st.error("❌ Please upload the MRF Template Excel file.")

    elif from_uploaded is None:
        st.error("❌ Please upload the From Solution Data Excel file.")

    elif rules_uploaded is None:
        st.error("❌ Please upload the Item Code Rules Excel file.")

    else:
        try:
            with st.spinner(
                "Generating MRF files and adding item codes... Please wait."
            ):
                created_files = create_mrf_files(
                    template_uploaded,
                    from_uploaded,
                    rules_uploaded
                )

            if not created_files:
                st.warning("⚠️ No MRF files were created.")

            else:
                st.success(
                    f"✅ Completed successfully! Total files created: "
                    f"{len(created_files)}"
                )

                st.divider()
                st.subheader("📁 Generated MRF Files")

                for file_path in created_files:
                    file_name = os.path.basename(file_path)

                    with open(file_path, "rb") as f:
                        file_data = f.read()

                    st.download_button(
                        label=f"⬇️ Download {file_name}",
                        data=file_data,
                        file_name=file_name,
                        mime=(
                            "application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet"
                        ),
                        use_container_width=True
                    )

                zip_buffer = io.BytesIO()

                with zipfile.ZipFile(
                    zip_buffer,
                    "w",
                    zipfile.ZIP_DEFLATED
                ) as zip_file:

                    for file_path in created_files:
                        zip_file.write(
                            file_path,
                            arcname=os.path.basename(file_path)
                        )

                zip_buffer.seek(0)

                st.divider()

                st.download_button(
                    label="📦 Download All MRF Files (ZIP)",
                    data=zip_buffer,
                    file_name="MRF_All_Files.zip",
                    mime="application/zip",
                    use_container_width=True
                )

        except Exception as e:
            st.error(
                "❌ An error occurred while generating the MRF files."
            )
            st.exception(e)
